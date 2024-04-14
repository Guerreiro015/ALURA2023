import os
os.system("cls" or None)

import openpyxl

# ABRIR UM ARQUIVO EXSITENTE

plan = openpyxl.load_workbook('dados.xlsx')

print(plan.sheetnames)

#selecionar aba que vai usar
planilha = plan.active
# ou escolhe a aba pelo nome
planilha = plan["plan1"]

#pecrorrer as linhas
lin = planilha.max_column # Numero de colunas
col = planilha.max_row # Numero de linhas

for linha in plan["plan1"].iter_rows(max_row = 2):
    #pecorrer as colunas
     for i in linha:
      print(i.value)
        
#     #Alterar valor das celulas
# for celula in plan.active["C"]:
#     if celula.value == "Vassora Sabiá":
#         linha = celula.row
#         plan.active[f"C{linha}"] = "Varredor"
#     print(celula.value)

tamanho = plan.active["a1:c20"]
for x,y,z in tamanho:
    print(x.value,y.value,z.value)


print(planilha.max_column) # Numero de colunas
print(planilha.max_row) # Numero de linhas




 
        
#plan.save("dados2.xlsx")
